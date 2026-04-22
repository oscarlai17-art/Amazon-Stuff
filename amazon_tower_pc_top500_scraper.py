"""
Amazon Tower PC Top 500 Scraper
================================
Scrapes up to 500 best-selling Tower PC products from Amazon by combining:
  1. BSR Top 100 (Best Sellers Rank pages 1-2)
  2. Search results sorted by popularity, paginated until 500 unique ASINs are collected

Output:
    amazon_tower_pc_top500.csv
    amazon_tower_pc_top500.xlsx
"""

import re
import time
import random
import pandas as pd
from playwright.sync_api import sync_playwright

# ── Config ─────────────────────────────────────────────────────────────────────
TARGET_PRODUCTS = 500

BSR_PAGES = [
    "https://www.amazon.com/Best-Sellers-Tower-Computers/zgbs/pc/13896597011/ref=zg_bs_pg_1?pg=1",
    "https://www.amazon.com/Best-Sellers-Tower-Computers/zgbs/pc/13896597011/ref=zg_bs_pg_2?pg=2",
]

SEARCH_URL_TEMPLATE = (
    "https://www.amazon.com/s?i=computers&bbn=13896597011"
    "&rh=n%3A13896597011&s=exact-aware-popularity-rank&page={page}"
)

MAX_SEARCH_PAGES = 25

OUTPUT_CSV   = "amazon_tower_pc_top500.csv"
OUTPUT_EXCEL = "amazon_tower_pc_top500.xlsx"

# ── GPU / Brand patterns ───────────────────────────────────────────────────────
GPU_PATTERNS = [
    r"RTX\s?\d{4}(?:\s?Ti|\s?Super)?",
    r"GTX\s?\d{4}(?:\s?Ti|\s?Super)?",
    r"RX\s?\d{4}(?:\s?XT|\s?XTX)?",
    r"Arc\s?[A-Z]\d{3,4}(?:\s?[A-Z])?",
    r"Radeon\s?(?:RX\s?)?\d{3,4}(?:M|XT|XTX)?",
    r"GeForce\s?(?:RTX|GTX)\s?\d{4}(?:\s?Ti|\s?Super)?",
    r"Intel\s?(?:Arc|UHD|Iris)",
    r"Vega\s?\d+",
    r"(?:RTX|GTX)\s?5\d{3}(?:\s?Ti)?",
    r"RX\s?9\d{3}(?:\s?XT)?",
]
GPU_REGEX = re.compile("|".join(GPU_PATTERNS), re.IGNORECASE)

BRAND_PATTERNS = [
    "Alienware", "HP", "OMEN", "CyberPowerPC", "iBuyPower", "MSI", "ASUS",
    "ROG", "Acer", "Predator", "Lenovo", "Legion", "Dell", "STGAubron",
    "Skytech", "NOVATECH", "KOTIN", "WIWB", "YAWYORE", "SKYESEV",
    "GMKtec", "GEEKOM", "Corsair", "Razer", "Maingear", "Velocity Micro",
]


def extract_gpu(text: str) -> str:
    match = GPU_REGEX.search(text or "")
    return match.group(0).strip() if match else "N/A"


def extract_brand(text: str, brand_from_amazon: str = "") -> str:
    if brand_from_amazon and brand_from_amazon.strip().lower() not in ("visit the store", "", "n/a"):
        return brand_from_amazon.strip()
    for b in BRAND_PATTERNS:
        if b.lower() in (text or "").lower():
            return b
    return "N/A"


def parse_price(price_str: str) -> str:
    if not price_str:
        return "N/A"
    price_str = price_str.strip().replace("\n", "").replace(",", "")
    match = re.search(r"\$[\d,]+\.?\d*", price_str)
    return match.group(0) if match else price_str[:20]


def extract_asin(url: str) -> str:
    match = re.search(r"/dp/([A-Z0-9]{10})", url or "")
    return match.group(1) if match else ""


def set_location(page):
    print("   Setting US delivery location (ZIP 10001 - New York)...")
    page.goto("https://www.amazon.com", wait_until="domcontentloaded", timeout=30000)
    time.sleep(3)
    try:
        location_btn = (
            page.query_selector("#nav-global-location-popover-link") or
            page.query_selector("#glow-ingress-block")
        )
        if location_btn:
            location_btn.click()
            time.sleep(2)
            zip_input = (
                page.query_selector("input[data-action='GLUXPostalUpdateAction']") or
                page.query_selector("input[id='GLUXZipUpdateInput']")
            )
            if zip_input:
                zip_input.click()
                zip_input.fill("")
                zip_input.type("10001", delay=80)
                time.sleep(1)
                apply_btn = (
                    page.query_selector("input.a-button-input[aria-labelledby='GLUXZipUpdate-announce']") or
                    page.query_selector("input[aria-labelledby='GLUXZipUpdate-announce']")
                )
                if apply_btn:
                    apply_btn.click()
                    time.sleep(2)
                done_btn = page.query_selector("button.a-button-text[name='glowDoneButton']")
                if done_btn:
                    done_btn.click()
                    time.sleep(2)
                print("   US location set to ZIP 10001 (New York)")
            else:
                print("   ZIP input not found — continuing anyway")
        else:
            print("   Location button not found — continuing anyway")
    except Exception as e:
        print(f"   Auto location failed ({e}) — continuing anyway")


def check_captcha(page):
    if "captcha" in page.content().lower() or "robot" in page.content().lower():
        print("   CAPTCHA detected! Solve it in the browser window (60s timeout)...")
        try:
            page.wait_for_function(
                "() => !document.body.innerText.toLowerCase().includes('captcha')",
                timeout=60000
            )
            print("   CAPTCHA solved! Continuing...")
            time.sleep(2)
            return True
        except Exception:
            print("   Timed out waiting for CAPTCHA.")
            return False
    return True


def slow_scroll(page, rounds=8):
    for _ in range(rounds):
        page.mouse.wheel(0, 600)
        time.sleep(random.uniform(0.8, 1.4))
    page.keyboard.press("Home")
    time.sleep(1)
    for _ in range(rounds):
        page.mouse.wheel(0, 600)
        time.sleep(0.4)
    time.sleep(1)


def scrape_bsr_cards(page, cards, seen_asins, products, rank_offset=0):
    rank = rank_offset
    for card in cards:
        rank += 1
        try:
            title_el = card.query_selector(
                "div._cDEzb_p13n-sc-css-line-clamp-3_g3dy1, "
                "span.a-size-base.a-color-base, "
                "div.p13n-sc-truncate-desktop-type2"
            )
            title = title_el.inner_text().strip() if title_el else "N/A"

            link_el = card.query_selector("a.a-link-normal")
            href = link_el.get_attribute("href") if link_el else ""
            full_url = f"https://www.amazon.com{href}" if href and href.startswith("/") else href or "N/A"
            asin = extract_asin(full_url)
            clean_url = f"https://www.amazon.com/dp/{asin}" if asin else full_url

            if asin and asin in seen_asins:
                continue
            if asin:
                seen_asins.add(asin)

            sell_el = card.query_selector(
                "span.a-color-price, "
                "span._cDEzb_p13n-sc-price_3mJ9Z, "
                "span.a-price > span.a-offscreen"
            )
            selling_price = parse_price(sell_el.inner_text() if sell_el else "")

            list_el = card.query_selector("span.a-text-price > span.a-offscreen")
            list_price = parse_price(list_el.inner_text() if list_el else "")
            if list_price == "N/A":
                list_price = selling_price

            brand_el = card.query_selector("span.a-size-small.a-color-base")
            brand_raw = brand_el.inner_text() if brand_el else ""
            brand = extract_brand(title, brand_raw)
            gpu = extract_gpu(title)

            products.append({
                "Rank":          len(products) + 1,
                "Source":        "BSR",
                "Title":         title,
                "Brand":         brand,
                "GPU":           gpu,
                "List Price":    list_price,
                "Selling Price": selling_price,
                "ASIN":          asin or "N/A",
                "URL":           clean_url,
            })
            print(f"   #{len(products):>3} [BSR] | {brand:<18} | {gpu:<20} | {selling_price:<10} | {title[:50]}")

        except Exception as e:
            print(f"   Error parsing BSR card: {e}")

    return products, seen_asins


def scrape_search_cards(page, seen_asins, products):
    cards = page.query_selector_all("div[data-component-type='s-search-result']")
    for card in cards:
        if len(products) >= TARGET_PRODUCTS:
            break
        try:
            title_el = card.query_selector("h2 span, h2 a span")
            title = title_el.inner_text().strip() if title_el else "N/A"

            link_el = card.query_selector("h2 a")
            href = link_el.get_attribute("href") if link_el else ""
            full_url = f"https://www.amazon.com{href}" if href and href.startswith("/") else href or "N/A"
            asin = extract_asin(full_url)
            clean_url = f"https://www.amazon.com/dp/{asin}" if asin else full_url

            if asin and asin in seen_asins:
                continue
            if asin:
                seen_asins.add(asin)

            price_el = card.query_selector("span.a-price > span.a-offscreen")
            selling_price = parse_price(price_el.inner_text() if price_el else "")

            list_price_el = card.query_selector("span.a-text-price > span.a-offscreen")
            list_price = parse_price(list_price_el.inner_text() if list_price_el else "")
            if list_price == "N/A":
                list_price = selling_price

            brand_el = card.query_selector("span.a-size-base-plus.a-color-base, h2 + div span")
            brand_raw = brand_el.inner_text() if brand_el else ""
            brand = extract_brand(title, brand_raw)
            gpu = extract_gpu(title)

            products.append({
                "Rank":          len(products) + 1,
                "Source":        "Search",
                "Title":         title,
                "Brand":         brand,
                "GPU":           gpu,
                "List Price":    list_price,
                "Selling Price": selling_price,
                "ASIN":          asin or "N/A",
                "URL":           clean_url,
            })
            print(f"   #{len(products):>3} [Search] | {brand:<18} | {gpu:<20} | {selling_price:<10} | {title[:45]}")

        except Exception as e:
            print(f"   Error parsing search card: {e}")

    return products, seen_asins


def scrape(target=TARGET_PRODUCTS) -> list[dict]:
    products = []
    seen_asins: set[str] = set()

    with sync_playwright() as p:
        browser = p.chromium.launch(channel="chrome", headless=False)
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/123.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1280, "height": 900},
            locale="en-US",
            timezone_id="America/New_York",
            geolocation={"latitude": 40.7128, "longitude": -74.0060},
            permissions=["geolocation"],
            extra_http_headers={
                "Accept-Language": "en-US,en;q=0.9",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            }
        )
        context.add_cookies([
            {"name": "i18n-prefs", "value": "USD", "domain": ".amazon.com", "path": "/"},
            {"name": "lc-main",   "value": "en_US", "domain": ".amazon.com", "path": "/"},
        ])
        page = context.new_page()

        # ── Step 1: Set location ───────────────────────────────────────────────
        set_location(page)

        # ── Step 2: BSR top 100 ────────────────────────────────────────────────
        print("\n--- Phase 1: BSR Top 100 ---")
        for url in BSR_PAGES:
            print(f"\nFetching BSR page: {url}")
            try:
                page.goto(url, wait_until="networkidle", timeout=60000)
                time.sleep(random.uniform(3, 5))
                slow_scroll(page)
                if not check_captcha(page):
                    break
                cards = page.query_selector_all("div.zg-grid-general-faceout, li.zg-item-immersion")
                print(f"   Found {len(cards)} cards")
                products, seen_asins = scrape_bsr_cards(page, cards, seen_asins, products)
            except Exception as e:
                print(f"Failed to load BSR page: {e}")
            wait = random.uniform(8, 14)
            print(f"   Waiting {wait:.1f}s...")
            time.sleep(wait)

        # ── Step 3: Search results sorted by popularity ────────────────────────
        print(f"\n--- Phase 2: Search results (target: {target} total) ---")
        for page_num in range(1, MAX_SEARCH_PAGES + 1):
            if len(products) >= target:
                break
            url = SEARCH_URL_TEMPLATE.format(page=page_num)
            print(f"\nFetching search page {page_num}: {url}")
            try:
                page.goto(url, wait_until="networkidle", timeout=60000)
                time.sleep(random.uniform(3, 5))
                slow_scroll(page)
                if not check_captcha(page):
                    break

                # Detect end of results
                no_results = page.query_selector("span.a-size-medium.a-color-base:has-text('No results')")
                if no_results:
                    print("   No more results — stopping.")
                    break

                before = len(products)
                products, seen_asins = scrape_search_cards(page, seen_asins, products)
                new_count = len(products) - before
                print(f"   Added {new_count} new products (total: {len(products)})")

                if new_count == 0:
                    print("   No new products found — Amazon may have stopped showing results.")
                    break

            except Exception as e:
                print(f"Failed to load search page {page_num}: {e}")
                break

            wait = random.uniform(8, 14)
            print(f"   Waiting {wait:.1f}s...")
            time.sleep(wait)

        browser.close()

    return products


def save_outputs(products: list[dict]):
    if not products:
        print("\nNo products scraped.")
        return

    df = pd.DataFrame(products)
    df.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")
    print(f"\nCSV saved: {OUTPUT_CSV}")

    with pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Top 500 Tower PCs")
        wb = writer.book
        ws = writer.sheets["Top 500 Tower PCs"]

        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="1F3864")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        url_col_idx = df.columns.get_loc("URL") + 1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            url_cell = row[url_col_idx - 1]
            if url_cell.value and str(url_cell.value).startswith("http"):
                url_cell.hyperlink = url_cell.value
                url_cell.value = "Open on Amazon"
                url_cell.font = Font(color="0070C0", underline="single")

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    print(f"Excel saved: {OUTPUT_EXCEL}")
    print(f"\nTotal products scraped: {len(products)}")


def main():
    print("=" * 60)
    print("  Amazon Tower PC Top 500 Scraper")
    print("=" * 60)
    print(f"\nTarget: {TARGET_PRODUCTS} unique products")
    print("Phase 1: BSR Top 100 | Phase 2: Search results\n")

    products = scrape()
    save_outputs(products)
    print("\nDone! Open the .xlsx file for the formatted report.")


if __name__ == "__main__":
    main()
