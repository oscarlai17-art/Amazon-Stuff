"""
Amazon Gaming PC BSR Scraper
=============================
Scrapes the Top 100 Best Sellers in Gaming PCs (Tower Computers) from Amazon.
Extracts: Rank, Title, Brand, GPU, List Price, Selling Price, URL

Requirements:
    pip install requests beautifulsoup4 playwright pandas openpyxl
    playwright install chromium

Usage:
    python amazon_gaming_pc_scraper.py

Output:
    amazon_gaming_pcs_top100.csv
    amazon_gaming_pcs_top100.xlsx
"""

import re
import time
import random
import pandas as pd
from playwright.sync_api import sync_playwright

# ── Config ─────────────────────────────────────────────────────────────────────
BSR_PAGES = [
    # Tower Computers BSR (150 products)
    "https://www.amazon.com/Best-Sellers-Tower-Computers/zgbs/pc/13896597011/ref=zg_bs_pg_1?pg=1",
    "https://www.amazon.com/Best-Sellers-Tower-Computers/zgbs/pc/13896597011/ref=zg_bs_pg_2?pg=2",
    "https://www.amazon.com/Best-Sellers-Tower-Computers/zgbs/pc/13896597011/ref=zg_bs_pg_3?pg=3",
    "https://www.amazon.com/Best-Sellers-Tower-Computers/zgbs/pc/13896597011/ref=zg_bs_pg_4?pg=4",
    "https://www.amazon.com/Best-Sellers-Tower-Computers/zgbs/pc/13896597011/ref=zg_bs_pg_5?pg=5",
    # Desktop Computers BSR (150 products)
    "https://www.amazon.com/Best-Sellers-Desktop-Computers/zgbs/electronics/565098/ref=zg_bs_pg_1?pg=1",
    "https://www.amazon.com/Best-Sellers-Desktop-Computers/zgbs/electronics/565098/ref=zg_bs_pg_2?pg=2",
    "https://www.amazon.com/Best-Sellers-Desktop-Computers/zgbs/electronics/565098/ref=zg_bs_pg_3?pg=3",
    "https://www.amazon.com/Best-Sellers-Desktop-Computers/zgbs/electronics/565098/ref=zg_bs_pg_4?pg=4",
    "https://www.amazon.com/Best-Sellers-Desktop-Computers/zgbs/electronics/565098/ref=zg_bs_pg_5?pg=5",
    # Gaming PCs search results sorted by best sellers (200+ products)
    "https://www.amazon.com/s?k=gaming+pc+desktop&rh=n%3A13896597011&s=exact-aware-popularity-rank&page=1",
    "https://www.amazon.com/s?k=gaming+pc+desktop&rh=n%3A13896597011&s=exact-aware-popularity-rank&page=2",
    "https://www.amazon.com/s?k=gaming+pc+desktop&rh=n%3A13896597011&s=exact-aware-popularity-rank&page=3",
    "https://www.amazon.com/s?k=gaming+pc+desktop&rh=n%3A13896597011&s=exact-aware-popularity-rank&page=4",
    "https://www.amazon.com/s?k=gaming+pc+desktop&rh=n%3A13896597011&s=exact-aware-popularity-rank&page=5",
]

OUTPUT_CSV   = "amazon_gaming_pcs_top100.csv"
OUTPUT_EXCEL = "amazon_gaming_pcs_top100.xlsx"

# GPU keyword patterns to extract from product titles/descriptions
GPU_PATTERNS = [
    r"RTX\s?\d{4}(?:\s?Ti|\s?Super)?",
    r"GTX\s?\d{4}(?:\s?Ti|\s?Super)?",
    r"RX\s?\d{4}(?:\s?XT|\s?XTX)?",
    r"Arc\s?[A-Z]\d{3,4}(?:\s?[A-Z])?",
    r"Radeon\s?(?:RX\s?)?\d{3,4}(?:M|XT|XTX)?",
    r"GeForce\s?(?:RTX|GTX)\s?\d{4}(?:\s?Ti|\s?Super)?",
    r"Intel\s?(?:Arc|UHD|Iris)",
    r"Vega\s?\d+",
    r"(?:RTX|GTX)\s?5\d{3}(?:\s?Ti)?",   # 5000-series
    r"RX\s?9\d{3}(?:\s?XT)?",            # RDNA 4
]

GPU_REGEX = re.compile("|".join(GPU_PATTERNS), re.IGNORECASE)

BRAND_PATTERNS = [
    "Alienware", "HP", "OMEN", "CyberPowerPC", "iBuyPower", "MSI", "ASUS",
    "ROG", "Acer", "Predator", "Lenovo", "Legion", "Dell", "STGAubron",
    "Skytech", "NOVATECH", "KOTIN", "WIWB", "YAWYORE", "SKYESEV",
    "GMKtec", "GEEKOM", "Corsair", "Razer", "Maingear", "Velocity Micro",
]

def extract_gpu(text: str) -> str:
    """Pull first GPU model found in text."""
    match = GPU_REGEX.search(text or "")
    return match.group(0).strip() if match else "N/A"

def extract_brand(text: str, brand_from_amazon: str) -> str:
    """Return Amazon brand field if available, else pattern-match from title."""
    if brand_from_amazon and brand_from_amazon.strip().lower() not in ("visit the store", "", "n/a"):
        return brand_from_amazon.strip()
    for b in BRAND_PATTERNS:
        if b.lower() in (text or "").lower():
            return b
    return "N/A"

def parse_price(price_str: str) -> str:
    """Normalize price string."""
    if not price_str:
        return "N/A"
    price_str = price_str.strip().replace("\n", "").replace(",", "")
    match = re.search(r"\$[\d,]+\.?\d*", price_str)
    return match.group(0) if match else price_str[:20]

def scrape_bsr(pages: list[str]) -> list[dict]:
    products = []
    rank_counter = 0

    with sync_playwright() as p:
        browser = p.chromium.launch(channel="chrome", headless=False)
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/123.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1280, "height": 900},
            locale="en-US",
            timezone_id="America/New_York",
            geolocation={"latitude": 40.7128, "longitude": -74.0060},
            permissions=["geolocation"],
            extra_http_headers={
                "Accept-Language": "en-US,en;q=0.9",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            }
        )
        context.add_cookies([
            {"name": "i18n-prefs", "value": "USD", "domain": ".amazon.com", "path": "/"},
            {"name": "lc-main", "value": "en_US", "domain": ".amazon.com", "path": "/"},
            {"name": "x-main", "value": "en_US", "domain": ".amazon.com", "path": "/"},
        ])
        page = context.new_page()

        # Open Amazon homepage first and set US zip code manually via location popup
        print("   🌎 Setting US delivery location (ZIP 10001 - New York)...")
        page.goto("https://www.amazon.com", wait_until="domcontentloaded", timeout=30000)
        time.sleep(3)

        try:
            location_btn = page.query_selector("#nav-global-location-popover-link")
            if location_btn:
                location_btn.click()
                time.sleep(2)
                zip_input = page.query_selector("input[data-action='GLUXPostalUpdateAction']")
                if zip_input:
                    zip_input.fill("10001")
                    time.sleep(1)
                    apply_btn = page.query_selector("input.a-button-input[aria-labelledby='GLUXZipUpdate-announce']")
                    if apply_btn:
                        apply_btn.click()
                        time.sleep(2)
                    done_btn = page.query_selector("button.a-button-text[name='glowDoneButton']")
                    if done_btn:
                        done_btn.click()
                        time.sleep(2)
                    print("   ✅ US location set to ZIP 10001 (New York)")
                else:
                    print("   ⚠️  ZIP input not found — please set location manually in the browser window then press Enter here")
                    input()
            else:
                print("   ⚠️  Location button not found — please set location manually in the browser window then press Enter here")
                input()
        except Exception as e:
            print(f"   ⚠️  Auto location failed ({e}) — please set location manually then press Enter here")
            input()

        for page_url in pages:
            print(f"\n📄 Fetching: {page_url}")
            try:
                page.goto(page_url, wait_until="networkidle", timeout=60000)
                time.sleep(random.uniform(3, 5))
                
                # Scroll down slowly to trigger lazy loading of all products
                for scroll in range(8):
                    page.mouse.wheel(0, 600)
                    time.sleep(random.uniform(0.8, 1.5))
                
                # Scroll back to top
                page.keyboard.press("Home")
                time.sleep(2)
                
                # Scroll down again fully to make sure everything is loaded
                for scroll in range(8):
                    page.mouse.wheel(0, 600)
                    time.sleep(0.5)
                
                time.sleep(2)   # polite delay

                # Check for CAPTCHA
                if "captcha" in page.content().lower() or "robot" in page.content().lower():
                    print("⚠️  CAPTCHA detected! Solve it in the browser window.")
                    print("    Waiting up to 60 seconds for you to solve it...")
                    try:
                        page.wait_for_function(
                            "() => !document.body.innerText.toLowerCase().includes('captcha')",
                            timeout=60000
                        )
                        print("    ✅ CAPTCHA solved! Continuing...")
                        time.sleep(2)
                    except:
                        print("    ❌ Timed out waiting for CAPTCHA. Try again.")
                        break

                # Each product card on BSR pages
                cards = page.query_selector_all("div.zg-grid-general-faceout, li.zg-item-immersion")
                print(f"   Found {len(cards)} product cards")

                for card in cards:
                    rank_counter += 1
                    try:
                        # Title
                        title_el = card.query_selector("div._cDEzb_p13n-sc-css-line-clamp-3_g3dy1, span.a-size-base.a-color-base, div.p13n-sc-truncate-desktop-type2")
                        title = title_el.inner_text().strip() if title_el else "N/A"

                        # URL / ASIN
                        link_el = card.query_selector("a.a-link-normal")
                        href = link_el.get_attribute("href") if link_el else ""
                        full_url = f"https://www.amazon.com{href}" if href and href.startswith("/") else href or "N/A"
                        # Clean URL to just product page
                        asin_match = re.search(r"/dp/([A-Z0-9]{10})", full_url)
                        clean_url = f"https://www.amazon.com/dp/{asin_match.group(1)}" if asin_match else full_url

                        # Prices
                        # Selling price
                        sell_el = card.query_selector("span.a-color-price, span._cDEzb_p13n-sc-price_3mJ9Z, span.a-price > span.a-offscreen")
                        selling_price = parse_price(sell_el.inner_text() if sell_el else "")

                        # List / was price
                        list_el = card.query_selector("span.a-text-price > span.a-offscreen, span.a-price.a-text-price span")
                        list_price = parse_price(list_el.inner_text() if list_el else "")
                        if list_price == "N/A":
                            list_price = selling_price  # no discount shown

                        # Brand (Amazon sometimes shows it)
                        brand_el = card.query_selector("span.a-size-small.a-color-base")
                        brand_raw = brand_el.inner_text() if brand_el else ""
                        brand = extract_brand(title, brand_raw)

                        # GPU from title
                        gpu = extract_gpu(title)

                        products.append({
                            "Rank":          rank_counter,
                            "Title":         title,
                            "Brand":         brand,
                            "GPU":           gpu,
                            "List Price":    list_price,
                            "Selling Price": selling_price,
                            "URL":           clean_url,
                        })

                        print(f"   #{rank_counter:>3} | {brand:<18} | {gpu:<22} | {selling_price:<10} | {title[:55]}")

                    except Exception as e:
                        print(f"   ⚠️  Error parsing card #{rank_counter}: {e}")
                        rank_counter -= 1  # don't count failed cards

            except Exception as e:
                print(f"❌ Failed to load page: {e}")

           # Polite pause between pages - longer to avoid detection
            wait = random.uniform(8, 15)
            print(f"   ⏳ Waiting {wait:.1f} seconds before next page...")
            time.sleep(wait)
            
            # Scroll around a bit to simulate human behavior
            page.mouse.wheel(0, random.randint(300, 700))
            time.sleep(random.uniform(1, 2))
            page.mouse.wheel(0, -random.randint(100, 300))
            time.sleep(random.uniform(1, 2))

        browser.close()

    return products


def save_outputs(products: list[dict]):
    if not products:
        print("\n❌ No products scraped. Check for CAPTCHAs or selector changes.")
        return

    df = pd.DataFrame(products)

    # CSV
    df.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")
    print(f"\n✅ CSV saved:   {OUTPUT_CSV}")

    # Excel with clickable hyperlinks
    with pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Top 100 Gaming PCs")
        wb = writer.book
        ws = writer.sheets["Top 100 Gaming PCs"]

        # Style header row
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="1F3864")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Make URL column clickable hyperlinks
        url_col_idx = df.columns.get_loc("URL") + 1  # 1-based
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            url_cell = row[url_col_idx - 1]
            if url_cell.value and url_cell.value.startswith("http"):
                url_cell.hyperlink = url_cell.value
                url_cell.value = "Open on Amazon →"
                url_cell.font = Font(color="0070C0", underline="single")

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    print(f"✅ Excel saved: {OUTPUT_EXCEL}")
    print(f"\n📊 Total products scraped: {len(products)}")


def main():
    print("=" * 60)
    print("  Amazon Gaming PC Top 100 BSR Scraper")
    print("=" * 60)
    print("\n⚙️  Starting scrape — this takes ~60–90 seconds...")
    print("   (polite delays between pages to avoid blocks)\n")

    products = scrape_bsr(BSR_PAGES)
    save_outputs(products)

    print("\n🏁 Done! Open the .xlsx file for a formatted report.")
    print("   Tip: If you hit CAPTCHAs, set headless=False in scrape_bsr()")
    print("        to see/solve them in a browser window.\n")


if __name__ == "__main__":
    main()
