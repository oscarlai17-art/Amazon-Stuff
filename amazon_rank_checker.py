"""
Amazon BSR Rank Checker
========================
Reads all ASINs from the latest "Last 2 days Sales" CSV and scrapes
their current Best Sellers Rank from each Amazon product page.

Output:
    Amazon-Data/Rankings/asin_rankings_YYYY-MM-DD.csv
    Amazon-Data/Rankings/asin_rankings_YYYY-MM-DD.xlsx

Usage:
    python amazon_rank_checker.py
"""

import os
import re
import glob
import time
import random
import pandas as pd
from datetime import date
from playwright.sync_api import sync_playwright

# ── Config ──────────────────────────────────────────────────────────────────
SALES_FOLDER  = "C:/Users/makep/Documents/Amazon-Data/Last 2 days Sales"
OUTPUT_FOLDER = "C:/Users/makep/Documents/Amazon-Data/Rankings"
TODAY         = date.today().strftime("%Y-%m-%d")
OUTPUT_CSV    = os.path.join(OUTPUT_FOLDER, f"asin_rankings_{TODAY}.csv")
OUTPUT_EXCEL  = os.path.join(OUTPUT_FOLDER, f"asin_rankings_{TODAY}.xlsx")


def load_asins() -> list[dict]:
    files = glob.glob(os.path.join(SALES_FOLDER, "*.csv"))
    if not files:
        raise FileNotFoundError(f"No CSV found in {SALES_FOLDER}")
    f = max(files, key=os.path.getmtime)
    print(f"Reading sales file: {os.path.basename(f)}")
    df = pd.read_csv(f, skiprows=1, encoding="utf-8-sig")
    # Only rows with actual sales
    df = df[pd.to_numeric(df["Ordered Units"], errors="coerce").fillna(0) > 0]
    return df[["ASIN", "Product Title", "Ordered Units", "Ordered Revenue"]].to_dict("records")


def extract_bsr(page) -> list[dict]:
    """Extract BSR entries using DOM selectors, falling back to regex."""
    results = []
    rank_re = re.compile(r"#([\d,]+)\s+in\s+(.+)")

    # Selectors that contain BSR text on Amazon product pages
    selectors = [
        "#detailBulletsWrapper_feature_div",
        "#productDetails_detailBullets_sections1",
        "#productDetails_db_sections",
        "#centerCol",
    ]

    text = ""
    for sel in selectors:
        el = page.query_selector(sel)
        if el:
            text = el.inner_text()
            if "Best Sellers Rank" in text or "Best Seller" in text:
                break

    if not text:
        text = page.inner_text("body")

    # Find the BSR block
    for line in text.splitlines():
        line = line.strip()
        m = rank_re.search(line)
        if m:
            rank = int(m.group(1).replace(",", ""))
            category = m.group(2).strip().split("(")[0].strip()
            if category and rank < 5_000_000:
                results.append({"rank": rank, "category": category})

    # Deduplicate
    seen, unique = set(), []
    for r in results:
        key = (r["rank"], r["category"])
        if key not in seen:
            seen.add(key)
            unique.append(r)
    return unique


def scrape_rankings(records: list[dict]) -> list[dict]:
    results = []

    with sync_playwright() as p:
        browser = p.chromium.launch(channel="chrome", headless=False)
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/123.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1280, "height": 900},
            locale="en-US",
            timezone_id="America/New_York",
        )
        context.add_cookies([
            {"name": "i18n-prefs", "value": "USD", "domain": ".amazon.com", "path": "/"},
            {"name": "lc-main",   "value": "en_US", "domain": ".amazon.com", "path": "/"},
        ])
        page = context.new_page()

        # Warm up on homepage
        print("Loading Amazon homepage...")
        page.goto("https://www.amazon.com", wait_until="domcontentloaded", timeout=30000)
        time.sleep(3)

        for i, rec in enumerate(records):
            asin  = rec["ASIN"]
            title = rec.get("Product Title", "")[:60]
            url   = f"https://www.amazon.com/dp/{asin}"

            print(f"\n[{i+1}/{len(records)}] {asin} — {title}")

            try:
                page.goto(url, wait_until="domcontentloaded", timeout=30000)

                # CAPTCHA check
                if "captcha" in page.content().lower() or "robot" in page.content().lower():
                    print("  ⚠️  CAPTCHA — solve in browser (60s timeout)...")
                    try:
                        page.wait_for_function(
                            "() => !document.body.innerText.toLowerCase().includes('captcha')",
                            timeout=60000
                        )
                        print("  ✅ CAPTCHA solved")
                    except Exception:
                        print("  ❌ CAPTCHA timeout — skipping")
                        results.append({**rec, "BSR Rank": "CAPTCHA", "BSR Category": ""})
                        continue

                # Scroll down to trigger lazy-loaded product details section
                for _ in range(6):
                    page.mouse.wheel(0, 600)
                    time.sleep(0.4)
                time.sleep(1.5)

                bsr_entries = extract_bsr(page)

                if bsr_entries:
                    # Primary = lowest rank number
                    primary = min(bsr_entries, key=lambda x: x["rank"])
                    all_ranks = " | ".join(
                        f"#{e['rank']:,} in {e['category']}" for e in bsr_entries
                    )
                    print(f"  ✅ Primary BSR: #{primary['rank']:,} in {primary['category']}")
                    results.append({
                        "ASIN":             asin,
                        "Product Title":    rec.get("Product Title", ""),
                        "Ordered Units":    rec.get("Ordered Units", ""),
                        "Ordered Revenue":  rec.get("Ordered Revenue", ""),
                        "Primary BSR Rank": primary["rank"],
                        "Primary Category": primary["category"],
                        "All Rankings":     all_ranks,
                    })
                else:
                    print("  ⚠️  No BSR found")
                    results.append({
                        "ASIN":             asin,
                        "Product Title":    rec.get("Product Title", ""),
                        "Ordered Units":    rec.get("Ordered Units", ""),
                        "Ordered Revenue":  rec.get("Ordered Revenue", ""),
                        "Primary BSR Rank": "N/A",
                        "Primary Category": "N/A",
                        "All Rankings":     "N/A",
                    })

            except Exception as e:
                print(f"  ❌ Error: {e}")
                results.append({
                    "ASIN":             asin,
                    "Product Title":    rec.get("Product Title", ""),
                    "Ordered Units":    rec.get("Ordered Units", ""),
                    "Ordered Revenue":  rec.get("Ordered Revenue", ""),
                    "Primary BSR Rank": "ERROR",
                    "Primary Category": str(e)[:80],
                    "All Rankings":     "",
                })

            # Polite delay between pages
            if i < len(records) - 1:
                wait = random.uniform(4, 8)
                time.sleep(wait)

        browser.close()

    return results


def save_outputs(results: list[dict]):
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    df = pd.DataFrame(results)
    df = df.sort_values("Ordered Units", ascending=False)

    df.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")
    print(f"\n✅ CSV saved: {OUTPUT_CSV}")

    with pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Rankings")
        wb = writer.book
        ws = writer.sheets["Rankings"]

        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="1F3864")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 80)

    print(f"✅ Excel saved: {OUTPUT_EXCEL}")
    print(f"\n📊 Checked {len(results)} products")
    print(df[["ASIN", "Primary BSR Rank", "Primary Category", "Ordered Units"]].to_string(index=False))


def main():
    print("=" * 60)
    print("  Amazon ASIN Rank Checker")
    print("=" * 60)

    records = load_asins()
    print(f"Found {len(records)} ASINs with sales to check\n")

    results = scrape_rankings(records)
    save_outputs(results)
    print("\n🏁 Done!")


if __name__ == "__main__":
    main()
